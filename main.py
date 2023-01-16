import os
import re
import sys
import datetime
import traceback
from PyQt5.QtWidgets import QFileDialog

from openpyxl import Workbook

from tubes_dialog import *


app = QtWidgets.QApplication(sys.argv)
MainWindow = QtWidgets.QMainWindow()
ui = Ui_MainWindow()
ui.setupUi(MainWindow)


def calc_tubes():
    try:
        dir = QFileDialog.getExistingDirectory()
        ui.label_2.setText(dir)
    except FileNotFoundError:
        return
    n = 0
    m = 0
    list_tubes, list_dates, list_files = [], [], []
    f = 0
    for top, dirs, files in os.walk(dir):
        for _ in files:
            f += 1
    ui.progressBar.setMaximum(f)
    for top, dirs, files in os.walk(dir):

        for file in files:
            n += 1
            ui.progressBar.setValue(n)
            if file.endswith('.qgd') and not file.lower().endswith('ms', 0, -4):
                qgd_file = file.split('.')[0]
                for i in qgd_file.split('_'):
                    if re.fullmatch(r'\d{6}', i):
                        tube = i
                        date = 0
                        if len(re.findall(r'\d{8}', qgd_file)) > 0:
                            date = datetime.datetime.strptime(re.findall(r'\d{8}', qgd_file)[0], '%d%m%Y')
                        if len(re.findall(r'\d{2}_\d{2}_\d{4}', qgd_file)) > 0:
                            date = datetime.datetime.strptime(re.findall(r'\d{2}_\d{2}_\d{4}', qgd_file)[0], '%d_%m_%Y')
                        if date:
                            if file not in list_files:
                                list_tubes.append(tube)
                                list_dates.append(date.strftime("%d.%m.%Y"))
                                list_files.append(file)
                                m += 1
                                ui.label.setText(f'Найдено файлов "*.qgd": {str(m)}')
                                print(m)

    print(len(list_tubes), list_tubes)
    print(len(list_dates), list_dates)
    print(len(list_files), list_files)

    wb = Workbook()
    ws = wb.active
    ws['A1'] = 'Трубка'
    ws['B1'] = 'Кол-во исп.'
    ws['C1'] = 'Даты'
    ws['D1'] = 'Файлы'
    row = 2

    for t in range(len(set(list_tubes))):
        indexes = [i for i, x in enumerate(list_tubes) if x == list_tubes[0]]
        ws[f'A{row}'] = list_tubes[indexes[0]]
        ws[f'B{row}'] = len(indexes)
        list_tube_dates, list_tube_files = [], []
        for i in indexes:
            list_tube_dates.append(list_dates[i])
            list_tube_files.append(list_files[i])
        ws[f'C{row}'] = ' / '.join(list_tube_dates)
        ws[f'D{row}'] = ' / '.join(list_tube_files)
        row += 1
        for i in reversed(indexes):
            del list_tubes[i]
            del list_dates[i]
            del list_files[i]

        print(len(list_tubes))
        ui.label.setText(f'Найдено трубок: {str(row - 2)}')


    fn = QFileDialog.getSaveFileName(caption="Сохранить результат в таблицу",
                                     directory=dir,
                                     filter="Excel Files (*.xlsx)")
    wb.save(fn[0])
    ui.label.setText(f'Сохранено в файл "{fn[0]}"')

MainWindow.show()


def log_uncaught_exceptions(ex_cls, ex, tb):
    """ Вывод ошибок программы """
    text = '{}: {}:\n'.format(ex_cls.__name__, ex)
    text += ''.join(traceback.format_tb(tb))
    print(text)
    QtWidgets.QMessageBox.critical(None, 'Error', text)
    sys.exit()


ui.pushButton.clicked.connect(calc_tubes)


sys.excepthook = log_uncaught_exceptions

sys.exit(app.exec_())